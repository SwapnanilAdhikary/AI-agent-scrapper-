import os
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def _flatten_analysis(analysis: dict) -> dict:
    """Flatten a single repo analysis into a flat dict suitable for a DataFrame row."""
    modules = analysis.get("modules", [])
    modules_str = "\n".join(
        f"- {m.get('name', '?')} ({m.get('category', '?')}): {m.get('description', '')}"
        for m in modules
    )
    modules_names = ", ".join(m.get("name", "?") for m in modules)

    strengths = analysis.get("strengths", [])
    weaknesses = analysis.get("weaknesses", [])

    return {
        "Rank": analysis.get("rank", "N/A"),
        "Repo Name": analysis.get("repo_name", ""),
        "URL": analysis.get("repo_url", ""),
        "Purpose": analysis.get("purpose", ""),
        "Tech Stack": analysis.get("tech_stack", ""),
        "Modules (Names)": modules_names,
        "Modules (Detail)": modules_str,
        "Module Count": len(modules),
        "Use Case Score": analysis.get("use_case_score", 0),
        "Code Quality Score": analysis.get("code_quality_score", 0),
        "Library Usage Score": analysis.get("library_usage_score", 0),
        "Overall Score": analysis.get("overall_score", 0),
        "Strengths": "\n".join(f"- {s}" for s in strengths) if strengths else "N/A",
        "Weaknesses": "\n".join(f"- {w}" for w in weaknesses) if weaknesses else "N/A",
        "Top 3": "Yes" if analysis.get("is_top_3") else "No",
        "Error": analysis.get("error", ""),
    }


def _style_excel(writer: pd.ExcelWriter, df_all: pd.DataFrame, df_top3: pd.DataFrame):
    """Apply formatting to the Excel workbook."""
    wb = writer.book

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for sheet_name, df in [("All Repos", df_all), ("Top 3 Rankings", df_top3)]:
        ws = wb[sheet_name]

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_length = max(
                len(str(col_name)),
                *(len(str(val).split("\n")[0]) for val in df[col_name]) if len(df) > 0 else [0],
            )
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"


def generate_reports(analyses: list[dict], output_dir: str) -> tuple[str, str]:
    """
    Generate CSV and Excel reports from the list of repo analyses.
    Returns (csv_path, excel_path).
    """
    os.makedirs(output_dir, exist_ok=True)

    rows = [_flatten_analysis(a) for a in analyses]
    df_all = pd.DataFrame(rows)

    df_all = df_all.sort_values("Rank", key=lambda x: pd.to_numeric(x, errors="coerce")).reset_index(drop=True)

    top3_rows = [r for r in rows if r["Top 3"] == "Yes"]
    df_top3 = pd.DataFrame(top3_rows) if top3_rows else pd.DataFrame(columns=df_all.columns)

    csv_path = os.path.join(output_dir, "github_analysis.csv")
    df_all.to_csv(csv_path, index=False, encoding="utf-8-sig")

    excel_path = os.path.join(output_dir, "github_analysis.xlsx")
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_all.to_excel(writer, sheet_name="All Repos", index=False)
        df_top3.to_excel(writer, sheet_name="Top 3 Rankings", index=False)
        _style_excel(writer, df_all, df_top3)

    return csv_path, excel_path
